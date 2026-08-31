# -*- coding: utf-8 -*-
"""
Deteccion ortografica rapida: extrae texto SIN abrir el documento en Calc/Writer.

Solo usa LibreOffice UNO para el servicio SpellChecker (mismas dicts es-GT / UA).
No marca, no sube rev_*.
"""
from __future__ import annotations

import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from spellcheck_core import (
    connect,
    find_unique_errors,
    make_locale,
)

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _errors_for_response(errores):
    out = []
    for e in errores or []:
        item = {
            "palabra": e.get("palabra", ""),
            "sugerencias": e.get("sugerencias") or [],
        }
        if e.get("llm_motivo"):
            item["llm_motivo"] = e.get("llm_motivo")
        out.append(item)
    return out


def extract_xlsx_text(path: str | Path) -> str:
    """Lectura rapida de .xlsx (openpyxl read_only)."""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    chunks: list[str] = []
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for val in row:
                    if val is None:
                        continue
                    if isinstance(val, str):
                        s = val.strip()
                    else:
                        s = str(val).strip()
                    if s:
                        chunks.append(s)
    finally:
        wb.close()
    return "\n".join(chunks)


def extract_docx_text(path: str | Path) -> str:
    """OOXML nativo: word/document.xml (sin python-docx)."""
    path = Path(path)
    parts: list[str] = []
    with zipfile.ZipFile(path, "r") as zf:
        # Documento principal
        try:
            root = ET.fromstring(zf.read("word/document.xml"))
            for node in root.iter(f"{W_NS}t"):
                if node.text:
                    parts.append(node.text)
        except KeyError:
            pass
        # Headers / footers (poco costo, a veces hay texto util)
        for name in zf.namelist():
            if not (
                name.startswith("word/header") or name.startswith("word/footer")
            ):
                continue
            if not name.endswith(".xml"):
                continue
            try:
                root = ET.fromstring(zf.read(name))
            except Exception:
                continue
            for node in root.iter(f"{W_NS}t"):
                if node.text:
                    parts.append(node.text)
    return "\n".join(parts)


def extract_pptx_text(path: str | Path) -> str:
    from spellcheck_pptx import extract_pptx_all_text

    return extract_pptx_all_text(path) or ""


def extract_pdf_text(path: str | Path) -> str:
    import fitz

    doc = fitz.open(str(path))
    try:
        return "\n".join(page.get_text("text") or "" for page in doc)
    finally:
        doc.close()


def extract_text_native(path: str | Path) -> tuple[str, str]:
    """
    Returns (texto, metodo_extraccion).
    Raises ValueError si el formato no tiene extractor rapido.
    """
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".xlsx":
        return extract_xlsx_text(path), "openpyxl_read_only"
    if ext == ".docx":
        return extract_docx_text(path), "ooxml_docx"
    if ext == ".pptx":
        return extract_pptx_text(path), "ooxml_pptx"
    if ext == ".pdf":
        return extract_pdf_text(path), "pymupdf"
    if ext in {".xls", ".doc", ".ppt"}:
        raise ValueError(
            f"Formato {ext} no soportado en fast-detect "
            "(usa .xlsx/.docx/.pptx/.pdf). Los binarios viejos requieren LO."
        )
    raise ValueError(f"Extension no soportada en fast-detect: {ext}")


def detect_fast(
    source_path: str | Path,
    *,
    llm_second_layer: bool = False,
) -> dict:
    """
    Extrae texto nativo + SpellChecker UNO (+ LLM opcional).
    No abre el documento Office en Calc/Writer.
    """
    t0 = time.perf_counter()
    source = Path(source_path)

    t_ext0 = time.perf_counter()
    text, metodo = extract_text_native(source)
    ms_extract = int((time.perf_counter() - t_ext0) * 1000)

    if not (text or "").strip():
        return {
            "ok": True,
            "archivo_original": source.name,
            "mensaje": "archivo sin contenido",
            "tiene_errores": False,
            "total_errores": 0,
            "errores": [],
            "capa_llm": bool(llm_second_layer),
            "solo_detectar": True,
            "extraccion": metodo,
            "ms_extraccion": ms_extract,
            "ms_spell": 0,
            "ms_llm": 0,
            "ms_total": int((time.perf_counter() - t0) * 1000),
        }

    t_sp0 = time.perf_counter()
    ctx = connect()
    smgr = ctx.ServiceManager
    spell = smgr.createInstanceWithContext(
        "com.sun.star.linguistic2.SpellChecker",
        ctx,
    )
    locale_es = make_locale("es", "GT")
    errores = find_unique_errors(text, spell, locale_es)
    ms_spell = int((time.perf_counter() - t_sp0) * 1000)
    candidatos_lo = len(errores)

    llm_meta = None
    ms_llm = 0
    if llm_second_layer and errores:
        t_llm0 = time.perf_counter()
        from llm_ortho_filter import filter_spelling_errors_with_llm

        errores, llm_meta = filter_spelling_errors_with_llm(errores)
        ms_llm = int((time.perf_counter() - t_llm0) * 1000)

    result = {
        "ok": True,
        "archivo_original": source.name,
        "archivo_rev": None,
        "tiene_errores": len(errores) > 0,
        "total_errores": len(errores),
        "errores": _errors_for_response(errores),
        "capa_llm": bool(llm_second_layer),
        "candidatos_libreoffice": candidatos_lo,
        "solo_detectar": True,
        "extraccion": metodo,
        "ms_extraccion": ms_extract,
        "ms_spell": ms_spell,
        "ms_llm": ms_llm,
        "ms_total": int((time.perf_counter() - t0) * 1000),
    }

    if llm_second_layer and llm_meta is not None:
        result["llm"] = {
            "aplicado": llm_meta.get("llm_aplicado"),
            "confirmados": llm_meta.get("confirmados"),
            "descartados": llm_meta.get("descartados"),
            "descartes": llm_meta.get("descartes"),
            "fallback": llm_meta.get("fallback"),
            "error": llm_meta.get("error"),
        }

    return result
